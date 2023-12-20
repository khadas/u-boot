#!/usr/bin/python3
# -*- coding:UTF-8 -*-

#  =====================================================================
#  @module:     Get all blx commits and save to csv file
#  @author:     Li Dongqing  (dongqing.li@amlogic.com)
#  @License:    Copyright (c) 2019 Amlogic, Inc. All rights reserved.
#  @Changes:
#               1. 2022.07.20 v0.1 create for stable branch update.
#  =====================================================================

# Import modules
try:
    import sys
    import os
    import re
    import json
    import time
    import argparse
    import subprocess
    import IPython
    import openpyxl
    from git.repo import Repo
    from openpyxl.styles import Font, Alignment
    from collections import OrderedDict
except Exception as e:
    print(e)
    exit('Please install modules, eg: pip3 install os')

# jason example: bl3.4.5-20220711-pre-ver.json
# {
#     "source_gits" : [
#         {"blType" : "bl2_sc2",         "gitPath" : "bl2/core_sc2",       "lastCommit" : "feebe5301418c038a06d45d0216c780ae9ea0033"},
#         {"blType" : "bl2_s4",          "gitPath" : "bl2/core_s4",        "lastCommit" : "817779a738e99b81081a31035ed784840cace44c"},
#         {"blType" : "bl2_ree",         "gitPath" : "bl2/ree",            "lastCommit" : "09d5c246638e95e2598a264b36da7f6ede7f6ea8"},
#         {"blType" : "bl2_tee",         "gitPath" : "bl2/tee",            "lastCommit" : "115e9fc38721c3564435f582875fc02908787b53"},
#         {"blType" : "bl30_aocpu",      "gitPath" : "bl30/src_ao",        "lastCommit" : "8c0b17692bd51f9c3311f8c51cd28bdf808a27a3"},
#         {"blType" : "bl31_1.3",        "gitPath" : "bl31_1.3/src",       "lastCommit" : "cf6108ce548d7ad3bfe268dedf801358664b6ead"},
#         {"blType" : "bl32_3.8",        "gitPath" : "bl32_3.8/src",       "lastCommit" : "c5ef42f2ce59304e2a4df7cf2dcbb12ab7ccefd1"},
#         {"blType" : "bl33",            "gitPath" : "bl33/v2019",         "lastCommit" : "f03ed9bc121114e9f31f1ee924d3adc176f13faa"},
#         {"blType" : "fip",             "gitPath" : "fip",                "lastCommit" : "d60ea7c9adfe8e537d0a11d2c2ce8e8097de5035"}
#     ]
# }

# bootloader trunk branch and remote info
pdPrefix  = "https://jira.amlogic.com/browse/"
scgitPrefix = "https://scgit.amlogic.com/plugins/gitiles/"
blSrcGits = [
    {"blType" : "bl2_sc2",                  "gitBranch" : "projects/sc2",             "gitRemote" : "firmware",         "upStream" : "bootloader/amlogic-advanced-bootloader/core/+/"},
    {"blType" : "bl2_s4",                   "gitBranch" : "projects/s4",              "gitRemote" : "firmware",         "upStream" : "bootloader/amlogic-advanced-bootloader/core/+/"},
    {"blType" : "bl2_s5",                   "gitBranch" : "projects/s5",              "gitRemote" : "firmware",         "upStream" : "bootloader/amlogic-advanced-bootloader/core/+/"},
    {"blType" : "bl2_ree",                  "gitBranch" : "projects/amlogic-dev",     "gitRemote" : "firmware",         "upStream" : "bootloader/amlogic-advanced-bootloader/ree/+/"},
    {"blType" : "bl2_tee",                  "gitBranch" : "projects/amlogic-dev",     "gitRemote" : "firmware",         "upStream" : "bootloader/amlogic-advanced-bootloader/tee/+/"},
    {"blType" : "bl30_aocpu",               "gitBranch" : "projects/amlogic-dev",     "gitRemote" : "firmware",         "upStream" : "firmware/aocpu/+/"},
    {"blType" : "rtos_arch",            "gitBranch" : "projects/amlogic-dev",     "gitRemote" : "origin",           "upStream" : "rtos_sdk/arch/riscv/+/"},
    {"blType" : "rtosboards",          "gitBranch" : "projects/amlogic-dev",     "gitRemote" : "origin",           "upStream" : "rtos_sdk/boards/riscv/+/"},
    {"blType" : "rtos_build",           "gitBranch" : "projects/amlogic-dev",     "gitRemote" : "origin",           "upStream" : "rtos_sdk/build/+/"},
    {"blType" : "rtos_driver_aocpu",    "gitBranch" : "projects/amlogic-dev",     "gitRemote" : "origin",           "upStream" : "rtos_sdk/drivers_aocpu/+/"},
    {"blType" : "rtos_freertos",        "gitBranch" : "projects/amlogic-dev",     "gitRemote" : "origin",           "upStream" : "rtos_sdk/freertos/+/"},
    {"blType" : "rtos_libc",            "gitBranch" : "projects/amlogic-dev",     "gitRemote" : "origin",           "upStream" : "rtos_sdk/libc/+/"},
    {"blType" : "rtos_products_aocpu",  "gitBranch" : "projects/amlogic-dev",     "gitRemote" : "origin",           "upStream" : "rtos_sdk/product/aocpu/+/"},
    {"blType" : "rtos_scripts",         "gitBranch" : "projects/amlogic-dev",     "gitRemote" : "origin",           "upStream" : "rtos_sdk/scripts/+/"},
    {"blType" : "rtos_soc",             "gitBranch" : "projects/amlogic-dev",     "gitRemote" : "origin",           "upStream" : "rtos_sdk/soc/riscv/+/"},
    {"blType" : "bl31_1.3",                 "gitBranch" : "amlogic-dev-1.3",          "gitRemote" : "firmware",         "upStream" : "ARM-software/arm-trusted-firmware/+/"},
    {"blType" : "bl32_3.8",                 "gitBranch" : "amlogic-dev-3.8.0",        "gitRemote" : "firmware",         "upStream" : "OP-TEE/optee_os/+/"},
    {"blType" : "bl32_3.18",                "gitBranch" : "amlogic-dev-3.18.0",        "gitRemote" : "firmware",        "upStream" : "OP-TEE/optee_os/+/"},
    {"blType" : "bl33",                     "gitBranch" : "amlogic-dev-2019",         "gitRemote" : "uboot",            "upStream" : "uboot/+/"},
    {"blType" : "fip",                      "gitBranch" : "amlogic-dev",              "gitRemote" : "fip",              "upStream" : "amlogic/tools/fip/+/"}
]

# the local csv file columns
csv_file_column = [
    {"ID" : "A",       "WIDTH" : 12,         "NAME" : "Index"},
    {"ID" : "B",       "WIDTH" : 45,         "NAME" : "Trunk Commit"},
    {"ID" : "C",       "WIDTH" : 30,         "NAME" : "Trunk Cl Link"},
    {"ID" : "D",       "WIDTH" : 20,         "NAME" : "Is Force Patch?"},
    {"ID" : "E",       "WIDTH" : 20,         "NAME" : "Is Secure Patch?"},
    {"ID" : "F",       "WIDTH" : 20,         "NAME" : "Reviewer"},
    {"ID" : "G",       "WIDTH" : 16,         "NAME" : "Related to other CL ?"},
    {"ID" : "H",       "WIDTH" : 20,         "NAME" : "QA Test Cases"},
    {"ID" : "I",       "WIDTH" : 16,         "NAME" : "QA Verify Result"},
    {"ID" : "J",       "WIDTH" : 20,         "NAME" : "New CL"}
]

# git update branch
def git_src_update(gitPath, gitRemote, gitBranch):
    local_path = os.path.join(gitPath)
    repo = Repo(local_path)

    try:
        repo.git.branch('-D', 'test')
    except:
        pass

    try:
        repo.git.checkout('-b', 'test')
    except:
        try:
            repo.git.checkout('test')
        except:
            pass
        pass

    try:
        repo.git.branch('-D', gitBranch)
    except:
        pass

    try:
        #repo.git.clean('-d', '-fx')
        repo.git.checkout('-t', 'remotes/' + gitRemote + '/' + gitBranch)
    except:
        try:
            repo.git.checkout(gitBranch)
        except:
            exit('Error: check out branch (%s / %s) failed!'%(gitRemote,gitBranch))
        pass

    try:
        repo.git.branch('-D', 'test')
    except:
        pass

    try:
        repo.git.fetch('--all')
    except:
        pass

    try:
        repo.git.reset('--hard', gitRemote + '/' + gitBranch)
    except:
        exit('Error: git reset branch (%s / %s) failed!'%(gitRemote,gitBranch))

    try:
        #repo.git.pull(gitRemote, gitBranch)
        repo.git.pull()
    except:
        exit('Error: git pull branch (%s) failed!'%gitBranch)

# git update branch
def get_bltype_branch_id(bltype, list):
    for i in range(len(list)):
        stream_dic = {"blType":None, "gitBranch":None, "gitRemote":None, "upStream":None}
        stream_dic = list[i]

        if str(bltype) == str(stream_dic['blType']):
            print(' >     Match the local bltype ID = ', i+1)
            return i

    else:
        return -1

# Prase json array to get commits info
def git_commits_to_src_link():
    print('\n[TRUNK LIST]:')

    # prase each repo
    for i in range(len(trunk_list)):
        print(' > [%d] gitPath: %-12s  lastCommit: %s'%(i+1, trunk_list[i]['gitPath'], trunk_list[i]['lastCommit']))

        if len(trunk_list[i]['lastCommit']) == 0:
            print(' >     lastCommit is NULL !')
            continue

        try:
            os.chdir(topdir + trunk_list[i]['gitPath'])
        except:
            exit('Error: NO such git path:', trunk_list[i]['gitPath'])

        # compare trunk list and local blSrcGits, find match id
        index = get_bltype_branch_id(trunk_list[i]['blType'], blSrcGits)
        if index < 0:
            exit("Error: NO found match blType:", trunk_list[i]['blType'])

        # update target branch
        git_src_update(os.getcwd(), blSrcGits[index]['gitRemote'], blSrcGits[index]['gitBranch'])

        # run git log format and produce commit list
        commit_list = git_cmt_parse(os.getcwd(), trunk_list[i]['lastCommit'], 'HEAD', 'TRUE')

        git_cmt_2_csv(csvfile, trunk_list[i]['blType'], commit_list, blSrcGits[index], i)

# Open json file, prase last commit
def prase_json_file():
    global trunk_list

    with open(jsonfile,'r') as load_f:
        try:
            json_array = json.load(load_f, object_pairs_hook=OrderedDict)
        except:
            exit('Error: Incorrect json format!')

        trunk_list = []
        for item in json_array['source_gits']:
            try:
                store_details = {"blType":None, "gitPath":None, "lastCommit":None}
                store_details['blType'] = item['blType']
                store_details['gitPath'] = item['gitPath']
                store_details['lastCommit'] = item['lastCommit']
                trunk_list.append(store_details)
            except:
                exit('Error: get trunk last commit failed.\n')

        return trunk_list

# output instance of str
def to_str(bytes_or_str):
    if isinstance(bytes_or_str, bytes):
        value = bytes_or_str.decode('utf-8')

    else:
        value = bytes_or_str

    return value

# run shell cmd return bytes
def bash_command(cmd):
    process = subprocess.Popen(cmd,
            shell=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE)

    #stdout, stderr = process.communicate()

    return process.stdout.read()

# get commit list info
def git_cmt_parse(gitPath, lastCommit, headCommit, isSrc):
    local_path = os.path.join(gitPath)

    repo = Repo(local_path)

    # run git log --format
    commit_log = repo.git.log('--pretty={"summary":"%s","commit":"%h","hash":"%H","author":"%ae","date":"%cd","pd":""}',
                            '--reverse', lastCommit + '...' + headCommit)

    try:
        log_list = commit_log.split("\n")
        #if debug_enable:
        #    print(' >    %s'%(log_list))
    except:
        pass
    #Remove invalid commits
    autosubmit = []
    log_list_len = len(log_list)
    for i in range(log_list_len):
        if "gerrit.autosubmit@amlogic.com" in log_list[i]:
            autosubmit.append(i)
    log_list = [log_list[i] for i in range(log_list_len) if (i not in autosubmit)]

    # deal with "Merge into" or "revert" commits
    for i in range(len(log_list)):
        try:
            log_list[i] = str(re.sub(r'Merge "', r'Merge <', str(log_list[i])))
            log_list[i] = str(re.sub(r'" into', r'> into', str(log_list[i])))
            log_list[i] = str(re.sub(r'Revert "', r'Revert <', str(log_list[i])))
            log_list[i] = str(re.sub(r'"",', r'>",', str(log_list[i])))
            log_list[i] = str(re.sub(r' "', r' <', str(log_list[i])))
            log_list[i] = str(re.sub(r'" ', r'> ', str(log_list[i])))

            if debug_enable:
                print(' >    [%d] %s'%(i,log_list[i]))
        except:
            pass
    # eval special special characters
    try:
        real_log_list = [eval(str(item)) for item in log_list]
    except:
        real_log_list = []
        if debug_enable:
            print(' >    eval(str(item)) ERROR!')
            print(' >    %s'%(log_list))
        pass
    # update real_log_list[i]['pd'] with JiraNo
    for j in range(len(real_log_list)):
        try:
            cmd = 'git log ' +  real_log_list[j]['commit'] + ' -1 | grep PD# | head -n 1'
            res = to_str(bash_command(cmd)).replace('\n', '')

            if res:
                real_log_list[j]['pd'] = res.split("PD#")[1]

            else:
                real_log_list[j]['pd'] = 'NULL'

            if debug_enable:
                print(' >    [%d] overwrite PD# = %s'%(j,real_log_list[j]['pd']))
        except:
            pass

    try:
        print(' >     Commit list max number of rows = ', len(real_log_list))
    except:
        pass

    return real_log_list

# save commit info to csv
def git_cmt_2_csv(csvfile, blType, commit_list, stream_dic, sheet_index):
    global alignment

    # Load a workbook object
    wb = openpyxl.load_workbook(csvfile)

    # Creat csv sheet named by blType
    title = re.sub(r'[?|$|.|!|/|*]', r'_', blType)
    sheet = wb.create_sheet(title, sheet_index)

    sheet.title = title

    # Set active sheet
    wb.active = sheet_index

    try:
        # Set aligment
        alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        for i in range(len(commit_list) + 1):

            for j in range(len(csv_file_column)):
                sheet.cell(row = i + 1, column = j + 1).alignment = alignment
    except:
        pass

    # set cell(1,1): value and font
    font = Font(size=11, bold=True)
    sheet.cell(row = 1, column = 1).value = blType + " " + csv_file_column[0]['NAME']
    sheet.cell(row = 1, column = 1).font = font

    # set row 1: value and font
    for i in range(1, len(csv_file_column)):
        sheet.cell(row = 1, column = i + 1).value = csv_file_column[i]['NAME']
        sheet.cell(row = 1, column = i + 1).font = font

    # set row 1: height
    # sheet.row_dimensions[1].height = 30

    # set row 2-n: trunk commit
    for i in range(len(commit_list)):
        try:
            # column 1: ID index
            sheet.cell(row = i + 2, column = 1).value = i + 1

            # column 2: Trunk Commit
            if commit_list[i]['pd'] == 'NULL':
                jira_pd = '\n'

            else:
                jira_pd = pdPrefix + commit_list[i]['pd'] + '\n'

            sheet.cell(row = i + 2, column = 2).value =     \
                                    jira_pd    + commit_list[i]['summary'] + '\n'\
                                    'Commit: ' + commit_list[i]['commit'] + '\n'\
                                    'Author: ' + commit_list[i]['author'] + '\n'\
                                    'Date:   ' + commit_list[i]['date']

            # column 3：Trunk Cl Link
            sheet.cell(row = i + 2, column = 3).value = scgitPrefix + \
                                                        stream_dic['upStream'] + \
                                                        commit_list[i]['hash']
        except:
            pass

    # set column A-I width
    for i in range(len(csv_file_column)):
        sheet.column_dimensions[csv_file_column[i]['ID']].width = csv_file_column[i]['WIDTH']

    # save the file
    wb.save(csvfile)

# Creat csv file to save cmt list
def create_csv_file(outdir, inputfile):
    global csvfile

    # get csv file full name
    localTime = time.strftime("_%Y%m%d_%H%M%S", time.localtime())

    base_name=os.path.basename(inputfile)
    file_name = os.path.splitext(base_name)[0]

    csvfile = str(topdir) + str(outdir) + "/" + file_name + str(localTime) + ".xlsx"

    # creat csv file, overwrite existing files
    wb = openpyxl.Workbook()

    wb.save(csvfile)

# get argv
def getOptions(args=sys.argv[1:]):
    global jsonfile
    global debug_enable

    # crate prase progress
    parser = argparse.ArgumentParser(description="Get all blx commits and save to csv file.")

    # add argv support
    parser.add_argument("-j", "--jsoncfg", required=True,  help="Your input json file.")
    parser.add_argument("-o", "--outDir",  required=True,  help="Your output .xlsx dir.")
    parser.add_argument("-v", "--verbose", required=False, help="Increase output verbosity.", action="store_true")
    parser.add_argument("-p", "--push",    required=False, help="Push csv file to confluence.", action="store_true")

    # prase argv
    options = parser.parse_args(args)

    # check debug on
    if options.verbose:
        debug_enable = 1
    else:
        debug_enable = 0

    # check whether the outdir is available
    jsonfile = str(options.jsoncfg)
    if not os.path.isdir(options.outDir):
        exit('Error: No such directory!')

    # creat new csv file
    create_csv_file(options.outDir, options.jsoncfg)
    print('\n[CONFIG INFO]:')
    print(' > Run in toplevel : ', topdir)
    print(' > Input json file : ', options.jsoncfg)
    print(' > Output xlsx dir : ', options.outDir)
    print(' > Output xlsx name: ', csvfile)

# Get the top-level directory which include .repo
def get_top_dir():
    global topdir

    pwd = os.getcwd()
    dirName = 'fip'

    if not os.path.exists(dirName):
        topdir = pwd + "/../"
    else:
        topdir = pwd + "/"

    return topdir

# Record summary info, and save to scv file
def record_in_summary_sheet():
    wb = openpyxl.load_workbook(csvfile)

    default_ws = 'Sheet'
    # active
    for s in range(len(wb.sheetnames)):
        if wb.sheetnames[s] == default_ws:
            break

    try:
        wb.active = s
        sheet = wb[default_ws]
        sheet.title = 'summary'
    except:
        exit('Error: NOT found sheet: %s'%default_ws)

    # set column width
    try:
        sheet.column_dimensions['A'].width = 10
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 15
    except:
        pass

    # row 1: summary head
    sheet.cell(row = 1, column = 1).value = 'ID'
    sheet.cell(row = 1, column = 2).value = 'title'
    sheet.cell(row = 1, column = 3).value = 'max_row'

    # row n: summary info
    for j in range(0,len(summary_list)):
        try:
            sheet.cell(row = j+2, column = 1).value = summary_list[j]['ID']
            sheet.cell(row = j+2, column = 2).value = summary_list[j]['title']
            sheet.cell(row = j+2, column = 3).value = summary_list[j]['max_row']

            if debug_enable:
                print('[%d] id:%d, title:%s, title:%2d'%(j+1,
                                        summary_list[j]['ID'],
                                        summary_list[j]['title'],
                                        summary_list[j]['max_row']))
        except:
            pass

    # Set aligment
    for i in range(0, len(summary_list) + 1):

        for j in range(0, 3):
            try:
                sheet.cell(row = i + 1, column = j + 1).alignment = alignment
            except:
                pass

    wb.save(csvfile)
    return

# Printf scv file , such as max list num
def summary_for_scv_sheets():
    global summary_list

    wb = openpyxl.load_workbook(csvfile)

    summary_list = []

    print('\n[SUMMARY SHEETS]:')
    print('=========================================')

    for i in range(0, len(wb.sheetnames)-1):
        try:
            sheet = wb[wb.sheetnames[i]]
            print('[%d] Sheet: %-12s    Max_Row: %2d'%(i+1, sheet.title, sheet.max_row-1))
            summary_details = {"ID":0, "title":None, "max_row":0}
            summary_details['ID'] = i+1
            summary_details['title'] = sheet.title
            summary_details['max_row'] = sheet.max_row-1
            summary_list.append(summary_details)
        except:
            pass

    print('=========================================\n')

    #print('summary_list: ', summary_list)

    wb.save(csvfile)
    return

# Main func
if __name__ == "__main__":
    # Set stderr with color
    from IPython.core.ultratb import ColorTB
    sys.excepthook = ColorTB()

    # get repo top dir
    get_top_dir()

    # Prase argv in / out
    getOptions(sys.argv[1:])

    # Prase json file to get last commit
    prase_json_file()

    # Get commits list and src link
    git_commits_to_src_link()

    # Output csv summary
    summary_for_scv_sheets()

    # Record in summary sheet
    record_in_summary_sheet()

    print('OUTPUT csv: ', os.path.basename(csvfile))
    exit('RUN OK !')

